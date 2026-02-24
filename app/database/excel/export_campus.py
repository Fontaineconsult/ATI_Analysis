import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.table import Table, TableStyleInfo


# Column definitions for campus tabs
COLUMNS = [
    'Goal',
    'Indicator ID',
    'Indicator',
    'Status',
    'Assigned Implementors',
    'Organizations',
    'Add Implementor',
]

ROW_KEYS = [
    'goal',
    'indicator_id',
    'indicator',
    'status',
    'implementors',
    'organizations',
    None,  # Add Implementor — left blank for dropdown selection
]

# Status column index (1-based) for color styling
STATUS_COL_IDX = COLUMNS.index('Status') + 1

# Styling constants
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
HEADER_ALIGNMENT = Alignment(horizontal="center", vertical="center", wrap_text=True)

WG_HEADER_FONT = Font(bold=True, color="FFFFFF", size=12)
WG_HEADER_FILL = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")

THIN_BORDER = Border(
    left=Side(style="thin", color="D9D9D9"),
    right=Side(style="thin", color="D9D9D9"),
    top=Side(style="thin", color="D9D9D9"),
    bottom=Side(style="thin", color="D9D9D9"),
)

# Status level colors: Red (Not Started) -> Green (Optimizing)
STATUS_FILLS = {
    "Not Started": PatternFill(start_color="FF4D4D", end_color="FF4D4D", fill_type="solid"),
    "Initiated":   PatternFill(start_color="FF9933", end_color="FF9933", fill_type="solid"),
    "Defined":     PatternFill(start_color="FFD966", end_color="FFD966", fill_type="solid"),
    "Established": PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"),
    "Managed":     PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid"),
    "Optimizing":  PatternFill(start_color="2E7D32", end_color="2E7D32", fill_type="solid"),
}
STATUS_FONTS = {
    "Not Started": Font(color="FFFFFF", bold=True),
    "Initiated":   Font(color="000000", bold=True),
    "Defined":     Font(color="000000", bold=True),
    "Established": Font(color="000000", bold=True),
    "Managed":     Font(color="FFFFFF", bold=True),
    "Optimizing":  Font(color="FFFFFF", bold=True),
}

DEFAULT_OUTPUT_DIR = os.path.join(os.path.dirname(__file__), "output")

# Counter for unique table names across sheets
_table_counter = 0


def _next_table_name(prefix):
    global _table_counter
    _table_counter += 1
    return f"{prefix}{_table_counter}"


def build_workbook(data: dict, output_dir: str = None) -> str:
    global _table_counter
    _table_counter = 0

    if output_dir is None:
        output_dir = DEFAULT_OUTPUT_DIR
    os.makedirs(output_dir, exist_ok=True)

    academic_year = data['academic_year']
    wb = Workbook()
    wb.remove(wb.active)

    # Reference sheets
    _build_persons_sheet(wb, data['all_persons'])
    _build_status_levels_sheet(wb, data.get('all_status_levels', []))

    # One tab per campus
    for campus in data['campuses']:
        _build_campus_sheet(
            wb,
            campus_label=campus['label'],
            campus_abbreviation=campus['abbreviation'],
            working_groups=campus['working_groups'],
            person_count=len(data['all_persons']),
        )

    filename = f"yse_campus_export_{academic_year}.xlsx"
    filepath = os.path.join(output_dir, filename)
    wb.save(filepath)
    return os.path.abspath(filepath)


def _build_persons_sheet(wb: Workbook, persons: list):
    """Create the Persons reference sheet with an Excel Table."""
    ws = wb.create_sheet("Persons")

    person_cols = ["Name", "Employee ID", "Organization", "Campus"]
    ws.append(person_cols)

    for p in sorted(persons, key=lambda x: (x['name'] or '')):
        ws.append([
            p['name'] or '',
            p['employee_id'] or '',
            p.get('organization', '') or '',
            p.get('campus', '') or '',
        ])

    # Create table
    if len(persons) > 0:
        end_row = len(persons) + 1
        table_ref = f"A1:{get_column_letter(len(person_cols))}{end_row}"
        table = Table(displayName="PersonsList", ref=table_ref)
        table.tableStyleInfo = TableStyleInfo(
            name="TableStyleMedium2", showFirstColumn=False,
            showLastColumn=False, showRowStripes=True, showColumnStripes=False
        )
        ws.add_table(table)

    _auto_fit_columns(ws)


def _build_status_levels_sheet(wb: Workbook, status_levels: list):
    """Create the Status Levels reference sheet with definitions."""
    ws = wb.create_sheet("Status Levels")

    headers = ["Status Level", "Value", "Procedures", "Documentation",
               "Documentation Evidence", "Resources"]
    ws.append(headers)

    for sl in status_levels:
        ws.append([
            sl.get('status_level', ''),
            sl.get('status_value', ''),
            sl.get('description_of_procedures', '') or '',
            sl.get('description_of_documentation', '') or '',
            sl.get('description_of_documentation_evidence', '') or '',
            sl.get('description_of_resources', '') or '',
        ])

    # Apply status colors to column A
    for row_idx in range(2, len(status_levels) + 2):
        cell = ws.cell(row=row_idx, column=1)
        status = cell.value
        if status in STATUS_FILLS:
            cell.fill = STATUS_FILLS[status]
            cell.font = STATUS_FONTS[status]
            cell.alignment = Alignment(horizontal="center", vertical="center")

    # Wrap text on description columns
    for row in ws.iter_rows(min_row=2, min_col=3, max_col=6):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    # Create table
    if len(status_levels) > 0:
        end_row = len(status_levels) + 1
        table_ref = f"A1:{get_column_letter(len(headers))}{end_row}"
        table = Table(displayName="StatusLevels", ref=table_ref)
        table.tableStyleInfo = TableStyleInfo(
            name="TableStyleMedium2", showFirstColumn=False,
            showLastColumn=False, showRowStripes=True, showColumnStripes=False
        )
        ws.add_table(table)

    _auto_fit_columns(ws, max_width=60)


def _build_campus_sheet(wb: Workbook, campus_label: str, campus_abbreviation: str,
                        working_groups: list, person_count: int):
    """
    Build a single campus tab containing all working groups.
    Each working group gets its own Excel Table for sorting/filtering.
    """
    ws = wb.create_sheet(campus_abbreviation.upper())

    # Title row
    ws.cell(row=1, column=1, value=f"{campus_label} — Assignment Tracker")
    ws.cell(row=1, column=1).font = Font(bold=True, size=14)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(COLUMNS))

    current_row = 3

    for wg in working_groups:
        # Working group header row (dark blue banner)
        ws.cell(row=current_row, column=1, value=wg['name'])
        for col_idx in range(1, len(COLUMNS) + 1):
            cell = ws.cell(row=current_row, column=col_idx)
            cell.font = WG_HEADER_FONT
            cell.fill = WG_HEADER_FILL
            cell.border = THIN_BORDER
        ws.merge_cells(
            start_row=current_row, start_column=1,
            end_row=current_row, end_column=len(COLUMNS)
        )
        current_row += 1

        if not wg['rows']:
            ws.cell(row=current_row, column=1, value="No data")
            ws.cell(row=current_row, column=1).font = Font(italic=True, color="999999")
            current_row += 2
            continue

        # Table header row
        table_header_row = current_row
        for col_idx, col_name in enumerate(COLUMNS, 1):
            cell = ws.cell(row=current_row, column=col_idx, value=col_name)
        current_row += 1

        # Write all data rows for this working group
        table_data_start = current_row
        data_row_numbers = []

        for row_data in wg['rows']:
            for col_idx, key in enumerate(ROW_KEYS, 1):
                if key is None:
                    value = ''
                else:
                    value = row_data.get(key) or ''
                cell = ws.cell(row=current_row, column=col_idx, value=value)
                cell.border = THIN_BORDER
                cell.alignment = Alignment(vertical="center", wrap_text=True)

                # Apply status color styling
                if col_idx == STATUS_COL_IDX and value in STATUS_FILLS:
                    cell.fill = STATUS_FILLS[value]
                    cell.font = STATUS_FONTS[value]
                    cell.alignment = Alignment(horizontal="center", vertical="center")

            data_row_numbers.append(current_row)
            current_row += 1

        # Create Excel Table for this working group
        table_end_row = current_row - 1
        end_col_letter = get_column_letter(len(COLUMNS))
        table_ref = f"A{table_header_row}:{end_col_letter}{table_end_row}"
        # Sanitize table name: no spaces, unique
        table_name = _next_table_name(
            campus_abbreviation.upper() + "_" + wg['name'].replace(" ", "").replace("/", "")[:15] + "_"
        )
        table = Table(displayName=table_name, ref=table_ref)
        table.tableStyleInfo = TableStyleInfo(
            name="TableStyleLight9", showFirstColumn=False,
            showLastColumn=False, showRowStripes=True, showColumnStripes=False
        )
        ws.add_table(table)

        # Data validation for "Add Implementor" column within this table
        add_impl_col = get_column_letter(COLUMNS.index('Add Implementor') + 1)
        if person_count > 0 and data_row_numbers:
            implementor_formula = f"Persons!$A$2:$A${person_count + 1}"
            dv_person = DataValidation(
                type="list",
                formula1=implementor_formula,
                allow_blank=True,
                showErrorMessage=True,
                errorTitle="Invalid Implementor",
                error="Please select a person from the Persons list.",
            )
            for r in data_row_numbers:
                dv_person.add(f"{add_impl_col}{r}")
            ws.add_data_validation(dv_person)

        # Blank row between working groups
        current_row += 1

    # Freeze below title
    ws.freeze_panes = "A3"

    _auto_fit_columns(ws)


def _auto_fit_columns(ws, min_width=10, max_width=50):
    """Set column widths based on maximum content length."""
    for col_cells in ws.columns:
        max_length = 0
        col_letter = get_column_letter(col_cells[0].column)
        for cell in col_cells:
            if cell.value is not None:
                try:
                    cell_len = len(str(cell.value))
                    if cell_len > max_length:
                        max_length = cell_len
                except Exception:
                    pass
        adjusted = max(min(max_length + 2, max_width), min_width)
        ws.column_dimensions[col_letter].width = adjusted
