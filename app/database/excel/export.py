import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation


# Column definitions for working group tabs
COLUMNS = [
    'Goal',
    'Indicator',
    'YSE',
    'Academic Year',
    'Implementor',
    'Employee ID',
    'Org Type',
    'Organization',
    'Campus',
]

# Row dict keys matching column order
ROW_KEYS = [
    'goal',
    'indicator',
    'yse',
    'academic_year',
    'implementor',
    'employee_id',
    'org_type',
    'organization',
    'campus',
]

# Styling constants
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
HEADER_ALIGNMENT = Alignment(horizontal="center", vertical="center", wrap_text=True)

GOAL_HEADER_FONT = Font(bold=True, size=11)
GOAL_HEADER_FILL = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")

# Alternating goal group shading
GOAL_SHADING = [
    PatternFill(start_color="F2F7FB", end_color="F2F7FB", fill_type="solid"),
    PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid"),
]

THIN_BORDER = Border(
    left=Side(style="thin", color="D9D9D9"),
    right=Side(style="thin", color="D9D9D9"),
    top=Side(style="thin", color="D9D9D9"),
    bottom=Side(style="thin", color="D9D9D9"),
)

DEFAULT_OUTPUT_DIR = os.path.join(os.path.dirname(__file__), "output")


def build_workbook(data: dict, output_dir: str = None) -> str:
    """
    Build an Excel workbook from the prepared export data.

    Args:
        data: dict from prepare_export_data() with keys:
              academic_year, working_groups, all_persons, all_campuses
        output_dir: directory for the output file; defaults to app/database/excel/output/

    Returns:
        Absolute path to the generated .xlsx file.
    """
    if output_dir is None:
        output_dir = DEFAULT_OUTPUT_DIR
    os.makedirs(output_dir, exist_ok=True)

    academic_year = data['academic_year']
    wb = Workbook()
    # Remove default sheet
    wb.remove(wb.active)

    # Build reference sheets first (needed for data validation ranges)
    _build_persons_sheet(wb, data['all_persons'])
    _build_campuses_sheet(wb, data['all_campuses'])

    # Build one tab per working group
    for wg in data['working_groups']:
        _build_working_group_sheet(
            wb, wg['name'], wg['rows'],
            len(data['all_persons']),
            len(data['all_campuses']),
        )

    # Save
    filename = f"yse_export_{academic_year}.xlsx"
    filepath = os.path.join(output_dir, filename)
    wb.save(filepath)
    return os.path.abspath(filepath)


def _build_persons_sheet(wb: Workbook, persons: list):
    """Create the Persons reference sheet used for data validation."""
    ws = wb.create_sheet("Persons")

    # Header
    ws.append(["Name", "Employee ID"])
    for cell in ws[1]:
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGNMENT

    # Data rows (sorted by name)
    for p in sorted(persons, key=lambda x: (x['name'] or '')):
        ws.append([p['name'] or '', p['employee_id'] or ''])

    _auto_fit_columns(ws)
    ws.sheet_state = 'visible'


def _build_campuses_sheet(wb: Workbook, campuses: list):
    """Create the Campuses reference sheet used for data validation."""
    ws = wb.create_sheet("Campuses")

    # Header
    ws.append(["Campus Name"])
    for cell in ws[1]:
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGNMENT

    # Data rows (sorted by name)
    for c in sorted(campuses, key=lambda x: (x['name'] or '')):
        ws.append([c['name'] or ''])

    _auto_fit_columns(ws)
    ws.sheet_state = 'visible'


def _build_working_group_sheet(wb: Workbook, wg_name: str, rows: list,
                                person_count: int, campus_count: int):
    """Build a single working-group tab with data, styling, and validation."""
    ws = wb.create_sheet(wg_name)

    # Write header row
    ws.append(COLUMNS)
    for col_idx, cell in enumerate(ws[1], 1):
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGNMENT
        cell.border = THIN_BORDER

    # Freeze the header row
    ws.freeze_panes = "A2"

    if not rows:
        _auto_fit_columns(ws)
        return

    # Group rows by goal to apply visual grouping
    goal_groups = _group_rows_by_goal(rows)

    current_row = 2  # row 1 is the header
    goal_color_index = 0
    data_row_numbers = []  # track actual data rows (not goal headers)

    for goal_name, goal_rows in goal_groups:
        # Write a bold goal header row spanning all columns
        ws.cell(row=current_row, column=1, value=goal_name)
        for col_idx in range(1, len(COLUMNS) + 1):
            cell = ws.cell(row=current_row, column=col_idx)
            cell.font = GOAL_HEADER_FONT
            cell.fill = GOAL_HEADER_FILL
            cell.border = THIN_BORDER
        # Merge the goal header across all columns
        ws.merge_cells(
            start_row=current_row, start_column=1,
            end_row=current_row, end_column=len(COLUMNS)
        )
        current_row += 1

        # Write data rows for this goal group
        fill = GOAL_SHADING[goal_color_index % len(GOAL_SHADING)]
        for row_data in goal_rows:
            for col_idx, key in enumerate(ROW_KEYS, 1):
                cell = ws.cell(row=current_row, column=col_idx,
                               value=row_data.get(key) or '')
                cell.fill = fill
                cell.border = THIN_BORDER
                cell.alignment = Alignment(vertical="center", wrap_text=True)
            data_row_numbers.append(current_row)
            current_row += 1

        goal_color_index += 1

    # Data validation — apply only to actual data rows, not goal header rows
    if person_count > 0 and data_row_numbers:
        implementor_formula = f"Persons!$A$2:$A${person_count + 1}"
        dv_person = DataValidation(
            type="list",
            formula1=implementor_formula,
            allow_blank=True,
            showErrorMessage=True,
            errorTitle="Invalid Implementor",
            error="Please select a person from the Persons list.",
        )
        for r in data_row_numbers:
            dv_person.add(f"E{r}")
        ws.add_data_validation(dv_person)

    if campus_count > 0 and data_row_numbers:
        campus_formula = f"Campuses!$A$2:$A${campus_count + 1}"
        dv_campus = DataValidation(
            type="list",
            formula1=campus_formula,
            allow_blank=True,
            showErrorMessage=True,
            errorTitle="Invalid Campus",
            error="Please select a campus from the Campuses list.",
        )
        for r in data_row_numbers:
            dv_campus.add(f"I{r}")
        ws.add_data_validation(dv_campus)

    _auto_fit_columns(ws)


def _group_rows_by_goal(rows: list) -> list:
    """
    Group rows by their 'goal' value, preserving order of first appearance.
    Returns list of (goal_name, [row_dicts]).
    """
    seen = {}
    order = []
    for row in rows:
        goal = row.get('goal') or 'No Goal'
        if goal not in seen:
            seen[goal] = []
            order.append(goal)
        seen[goal].append(row)
    return [(g, seen[g]) for g in order]


def _auto_fit_columns(ws, min_width=10, max_width=50):
    """Set column widths based on maximum content length."""
    for col_cells in ws.columns:
        max_length = 0
        col_letter = get_column_letter(col_cells[0].column)
        for cell in col_cells:
            if cell.value is not None:
                # Account for merged cells
                try:
                    cell_len = len(str(cell.value))
                    if cell_len > max_length:
                        max_length = cell_len
                except Exception:
                    pass
        adjusted = max(min(max_length + 2, max_width), min_width)
        ws.column_dimensions[col_letter].width = adjusted
