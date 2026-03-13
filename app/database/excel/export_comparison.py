import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo


# Campus display order and labels
CAMPUS_ORDER = ["ssu", "sfsu", "csueb"]
CAMPUS_LABELS = {"ssu": "SSU", "sfsu": "SFSU", "csueb": "CSUEB"}

# ============================================================
# Campus color palette — reusable across the workbook
# Full colors for text/accents, light versions for cell backgrounds
# ============================================================
CAMPUS_COLORS = {
    "ssu": {
        "full": "1B2A4A",   # Navy Blue
        "light": "D6DCE8",  # Light navy tint
    },
    "sfsu": {
        "full": "5B2C82",   # Purple
        "light": "E2D4F0",  # Light purple tint
    },
    "csueb": {
        "full": "B8860B",   # Gold (dark goldenrod)
        "light": "F5EDCF",  # Light gold tint
    },
}

# Pre-built fill objects for each campus
CAMPUS_BG_FILLS = {
    abbrev: PatternFill(start_color=colors["light"], end_color=colors["light"], fill_type="solid")
    for abbrev, colors in CAMPUS_COLORS.items()
}
CAMPUS_HEADER_FILLS = {
    abbrev: PatternFill(start_color=colors["full"], end_color=colors["full"], fill_type="solid")
    for abbrev, colors in CAMPUS_COLORS.items()
}
CAMPUS_HEADER_FONTS = {
    abbrev: Font(bold=True, color="FFFFFF", size=11)
    for abbrev in CAMPUS_COLORS
}

# Column layout: 11 columns total
COLUMNS = [
    "#",
    "Success Indicator",
    "SSU Status", "SFSU Status", "CSUEB Status",
    "SSU Impl & Notes", "SFSU Impl & Notes", "CSUEB Impl & Notes",
    "SSU People & Orgs", "SFSU People & Orgs", "CSUEB People & Orgs",
]

NUM_COLS = len(COLUMNS)

# Section boundaries (1-based column indices)
STATUS_COL_START = 3
STATUS_COL_END = 5
IMPL_COL_START = 6
IMPL_COL_END = 8
PEOPLE_COL_START = 9
PEOPLE_COL_END = 11

SECTION_HEADERS = [
    (STATUS_COL_START, STATUS_COL_END, "Status"),
    (IMPL_COL_START, IMPL_COL_END, "Implementations & Notes"),
    (PEOPLE_COL_START, PEOPLE_COL_END, "People & Organizations"),
]

# Map each column (1-based) to its campus index (0=ssu, 1=sfsu, 2=csueb) or None
COL_CAMPUS_INDEX = {
    3: 0, 4: 1, 5: 2,     # Status
    6: 0, 7: 1, 8: 2,     # Impl & Notes
    9: 0, 10: 1, 11: 2,   # People & Orgs
}

# Columns that form the RIGHT edge of a campus group (black right border)
# SSU cols: 3,6,9  SFSU cols: 4,7,10  CSUEB cols: 5,8,11
# Black separator after each campus's last column in a section AND between sections
SECTION_RIGHT_EDGE_COLS = {5, 8, 11}  # right edge of each 3-col section
CAMPUS_GROUP_RIGHT_EDGE_COLS = {3, 4, 6, 7, 9, 10}  # internal campus boundaries

# Border helpers
_BLACK_SIDE = Side(style="medium", color="000000")   # Thick vertical section separators
_GREY_SIDE = Side(style="thin", color="D9D9D9")      # Light borders for non-campus cells
_ROW_SIDE = Side(style="thin", color="808080")        # Visible row separators in campus areas

# Styling constants
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
HEADER_ALIGNMENT = Alignment(horizontal="center", vertical="center", wrap_text=True)

GOAL_BANNER_FONT = Font(bold=True, color="FFFFFF", size=12)
GOAL_BANNER_FILL = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")

SECTION_HEADER_FONT = Font(bold=True, color="FFFFFF", size=10)
SECTION_HEADER_FILL = PatternFill(start_color="5B9BD5", end_color="5B9BD5", fill_type="solid")

THIN_BORDER = Border(
    left=_GREY_SIDE, right=_GREY_SIDE,
    top=_GREY_SIDE, bottom=_GREY_SIDE,
)

SPACER_FILL = PatternFill(start_color="F5F5F5", end_color="F5F5F5", fill_type="solid")

STATUS_FILLS = {
    "Not Started": PatternFill(start_color="FF4D4D", end_color="FF4D4D", fill_type="solid"),
    "Initiated":   PatternFill(start_color="FF9933", end_color="FF9933", fill_type="solid"),
    "Defined":     PatternFill(start_color="FFD966", end_color="FFD966", fill_type="solid"),
    "Established": PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"),
    "Managed":     PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid"),
    "Optimizing":  PatternFill(start_color="2E7D32", end_color="2E7D32", fill_type="solid"),
}
STATUS_FONTS = {
    "Not Started": Font(color="FFFFFF", bold=True),
    "Initiated":   Font(color="000000", bold=True),
    "Defined":     Font(color="000000", bold=True),
    "Established": Font(color="000000", bold=True),
    "Managed":     Font(color="FFFFFF", bold=True),
    "Optimizing":  Font(color="FFFFFF", bold=True),
}


def _campus_border(col_idx):
    """Return a Border with medium-weight black verticals at section boundaries
    and visible grey horizontals for row separation in campus areas."""
    right = _BLACK_SIDE if col_idx in SECTION_RIGHT_EDGE_COLS else _ROW_SIDE
    left = _BLACK_SIDE if (col_idx - 1) in SECTION_RIGHT_EDGE_COLS or col_idx == 3 else _ROW_SIDE
    return Border(left=left, right=right, top=_ROW_SIDE, bottom=_ROW_SIDE)

DEFAULT_OUTPUT_DIR = os.path.join(os.path.dirname(__file__), "output")

_table_counter = 0


def _next_table_name(prefix):
    global _table_counter
    _table_counter += 1
    return f"{prefix}{_table_counter}"


def build_workbook(data: dict, output_dir: str = None) -> str:
    global _table_counter
    _table_counter = 0

    if output_dir is None:
        output_dir = DEFAULT_OUTPUT_DIR
    os.makedirs(output_dir, exist_ok=True)

    academic_year = data["academic_year"]
    wb = Workbook()
    wb.remove(wb.active)

    # Tab 1: Stats dashboard
    _build_stats_sheet(wb, data["working_groups"])

    # Tab 2: Status Levels reference
    _build_status_levels_sheet(wb, data.get("all_status_levels", []))

    # Tab 3: People reference
    _build_persons_sheet(wb, data["all_persons"])

    # Tab 4: Documents reference with hyperlinks
    _build_documents_sheet(wb, data.get("all_documents", []))

    # Tabs 5-7: Working group comparison sheets
    for wg in data["working_groups"]:
        _build_comparison_sheet(wb, wg["name"], wg["goals"])

    filename = f"campus_comparison_{academic_year}.xlsx"
    filepath = os.path.join(output_dir, filename)
    wb.save(filepath)
    return os.path.abspath(filepath)


def _build_persons_sheet(wb: Workbook, persons: list):
    """Create the Persons reference sheet with an Excel Table."""
    ws = wb.create_sheet("People")

    person_cols = ["Name", "Employee ID", "Organization", "Campus"]
    ws.append(person_cols)

    for p in sorted(persons, key=lambda x: (x['name'] or '')):
        ws.append([
            p['name'] or '',
            p['employee_id'] or '',
            p.get('organization', '') or '',
            p.get('campus', '') or '',
        ])

    if len(persons) > 0:
        end_row = len(persons) + 1
        table_ref = f"A1:{get_column_letter(len(person_cols))}{end_row}"
        table = Table(displayName="PersonsList", ref=table_ref)
        table.tableStyleInfo = TableStyleInfo(
            name="TableStyleMedium2", showFirstColumn=False,
            showLastColumn=False, showRowStripes=True, showColumnStripes=False
        )
        ws.add_table(table)

    _auto_fit_columns(ws)


def _build_status_levels_sheet(wb: Workbook, status_levels: list):
    """Create the Status Levels reference sheet with definitions."""
    ws = wb.create_sheet("Status")

    headers = ["Status Level", "Value", "Procedures", "Documentation",
               "Documentation Evidence", "Resources"]
    ws.append(headers)

    for sl in status_levels:
        ws.append([
            sl.get('status_level', ''),
            sl.get('status_value', ''),
            sl.get('description_of_procedures', '') or '',
            sl.get('description_of_documentation', '') or '',
            sl.get('description_of_documentation_evidence', '') or '',
            sl.get('description_of_resources', '') or '',
        ])

    for row_idx in range(2, len(status_levels) + 2):
        cell = ws.cell(row=row_idx, column=1)
        status = cell.value
        if status in STATUS_FILLS:
            cell.fill = STATUS_FILLS[status]
            cell.font = STATUS_FONTS[status]
            cell.alignment = Alignment(horizontal="center", vertical="center")

    for row in ws.iter_rows(min_row=2, min_col=3, max_col=6):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    if len(status_levels) > 0:
        end_row = len(status_levels) + 1
        table_ref = f"A1:{get_column_letter(len(headers))}{end_row}"
        table = Table(displayName="StatusLevels", ref=table_ref)
        table.tableStyleInfo = TableStyleInfo(
            name="TableStyleMedium2", showFirstColumn=False,
            showLastColumn=False, showRowStripes=True, showColumnStripes=False
        )
        ws.add_table(table)

    _auto_fit_columns(ws, max_width=60)


def _build_documents_sheet(wb: Workbook, documents: list):
    """Create the Documents reference sheet with clickable hyperlinks."""
    ws = wb.create_sheet("Documents")

    doc_cols = ["Type", "Name", "URL", "Implementation", "Campus", "Indicators"]
    ws.append(doc_cols)

    link_font = Font(color="0563C1", underline="single")

    for idx, doc in enumerate(documents, start=2):
        ws.cell(row=idx, column=1, value=doc['type'])
        name_cell = ws.cell(row=idx, column=2, value=doc['name'])
        url = doc.get('url', '')
        ws.cell(row=idx, column=3, value=url)

        # Make the name cell a clickable hyperlink if a URL exists
        if url:
            name_cell.hyperlink = url
            name_cell.font = link_font

        impl_label = f"[{doc['implementation_type']}] {doc['implementation_title']}" if doc['implementation_type'] else ''
        ws.cell(row=idx, column=4, value=impl_label)
        ws.cell(row=idx, column=5, value=doc.get('campus', '').upper())
        ws.cell(row=idx, column=6, value=doc.get('indicators', ''))

    if len(documents) > 0:
        end_row = len(documents) + 1
        table_ref = f"A1:{get_column_letter(len(doc_cols))}{end_row}"
        table = Table(displayName="DocumentsList", ref=table_ref)
        table.tableStyleInfo = TableStyleInfo(
            name="TableStyleMedium2", showFirstColumn=False,
            showLastColumn=False, showRowStripes=True, showColumnStripes=False
        )
        ws.add_table(table)

    _auto_fit_columns(ws, max_width=60)


def _build_stats_sheet(wb: Workbook, working_groups: list):
    """Build a Stats dashboard with status distribution and unassigned YSE."""
    ws = wb.create_sheet("Stats")

    # ── Collect stats from working group data ──
    # Status counts per campus
    status_counts = {a: {} for a in CAMPUS_ORDER}
    campus_totals = {a: 0 for a in CAMPUS_ORDER}
    # Unassigned: indicators where a campus has no people
    unassigned = []

    for wg in working_groups:
        for goal in wg["goals"]:
            for ind in goal["indicators"]:
                missing_campuses = []
                for abbrev in CAMPUS_ORDER:
                    campus_data = ind["campuses"][abbrev]
                    status = campus_data["status"] or "(No Status)"
                    status_counts[abbrev][status] = status_counts[abbrev].get(status, 0) + 1
                    campus_totals[abbrev] += 1
                    if not campus_data["people_and_orgs"].strip():
                        missing_campuses.append(abbrev)
                if missing_campuses:
                    unassigned.append({
                        "indicator_id": ind["indicator_id"],
                        "indicator_name": ind["indicator_name"],
                        "working_group": wg["name"],
                        "missing": missing_campuses,
                    })

    # Collect all status levels across campuses, ordered by the standard list
    status_order = list(STATUS_FILLS.keys()) + ["(No Status)"]
    all_statuses = set()
    for counts in status_counts.values():
        all_statuses.update(counts.keys())
    # Keep known statuses in order, then any extras
    ordered_statuses = [s for s in status_order if s in all_statuses]
    ordered_statuses += sorted(all_statuses - set(ordered_statuses))

    # ── Section 1: Status Level Distribution by Campus ──
    ws.cell(row=1, column=1, value="Status Level Distribution by Campus")
    ws.cell(row=1, column=1).font = Font(bold=True, size=14)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=7)

    dist_headers = ["Status Level",
                    "SSU Count", "SSU %",
                    "SFSU Count", "SFSU %",
                    "CSUEB Count", "CSUEB %"]
    for col_idx, h in enumerate(dist_headers, 1):
        cell = ws.cell(row=3, column=col_idx, value=h)
        cell.font = HEADER_FONT
        cell.alignment = HEADER_ALIGNMENT
        # Color the campus header pairs
        if col_idx in (2, 3):
            cell.fill = CAMPUS_HEADER_FILLS["ssu"]
            cell.font = CAMPUS_HEADER_FONTS["ssu"]
        elif col_idx in (4, 5):
            cell.fill = CAMPUS_HEADER_FILLS["sfsu"]
            cell.font = CAMPUS_HEADER_FONTS["sfsu"]
        elif col_idx in (6, 7):
            cell.fill = CAMPUS_HEADER_FILLS["csueb"]
            cell.font = CAMPUS_HEADER_FONTS["csueb"]
        else:
            cell.fill = HEADER_FILL

    for row_offset, status in enumerate(ordered_statuses):
        row = 4 + row_offset
        # Status label with color fill
        cell = ws.cell(row=row, column=1, value=status)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        if status in STATUS_FILLS:
            cell.fill = STATUS_FILLS[status]
            cell.font = STATUS_FONTS[status]

        for i, abbrev in enumerate(CAMPUS_ORDER):
            count = status_counts[abbrev].get(status, 0)
            total = campus_totals[abbrev]
            pct = (count / total * 100) if total > 0 else 0
            count_col = 2 + i * 2
            pct_col = 3 + i * 2
            ws.cell(row=row, column=count_col, value=count).alignment = Alignment(horizontal="center")
            pct_cell = ws.cell(row=row, column=pct_col, value=pct / 100)
            pct_cell.number_format = '0.0%'
            pct_cell.alignment = Alignment(horizontal="center")
            # Light campus background
            ws.cell(row=row, column=count_col).fill = CAMPUS_BG_FILLS[abbrev]
            pct_cell.fill = CAMPUS_BG_FILLS[abbrev]

    dist_end_row = 3 + len(ordered_statuses)

    # Add totals row
    totals_row = dist_end_row + 1
    ws.cell(row=totals_row, column=1, value="Total").font = Font(bold=True)
    for i, abbrev in enumerate(CAMPUS_ORDER):
        count_col = 2 + i * 2
        pct_col = 3 + i * 2
        ws.cell(row=totals_row, column=count_col, value=campus_totals[abbrev]).font = Font(bold=True)
        ws.cell(row=totals_row, column=count_col).alignment = Alignment(horizontal="center")
        pct_cell = ws.cell(row=totals_row, column=pct_col, value=1.0)
        pct_cell.number_format = '0.0%'
        pct_cell.font = Font(bold=True)
        pct_cell.alignment = Alignment(horizontal="center")

    # Excel Table for distribution
    table_ref = f"A3:G{totals_row}"
    table = Table(displayName="StatusDistribution", ref=table_ref)
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2", showFirstColumn=True,
        showLastColumn=False, showRowStripes=True, showColumnStripes=False,
    )
    ws.add_table(table)

    # ── Section 2: YSE with No People Assigned ──
    section2_start = totals_row + 3
    ws.cell(row=section2_start, column=1, value="Success Indicators with No People Assigned")
    ws.cell(row=section2_start, column=1).font = Font(bold=True, size=14)
    ws.merge_cells(start_row=section2_start, start_column=1,
                   end_row=section2_start, end_column=6)

    ua_header_row = section2_start + 1
    ua_headers = ["#", "Success Indicator", "Working Group", "SSU", "SFSU", "CSUEB"]
    for col_idx, h in enumerate(ua_headers, 1):
        cell = ws.cell(row=ua_header_row, column=col_idx, value=h)
        cell.font = HEADER_FONT
        cell.alignment = HEADER_ALIGNMENT
        if col_idx == 4:
            cell.fill = CAMPUS_HEADER_FILLS["ssu"]
            cell.font = CAMPUS_HEADER_FONTS["ssu"]
        elif col_idx == 5:
            cell.fill = CAMPUS_HEADER_FILLS["sfsu"]
            cell.font = CAMPUS_HEADER_FONTS["sfsu"]
        elif col_idx == 6:
            cell.fill = CAMPUS_HEADER_FILLS["csueb"]
            cell.font = CAMPUS_HEADER_FONTS["csueb"]
        else:
            cell.fill = HEADER_FILL

    missing_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    missing_font = Font(color="9C0006", bold=True)

    if unassigned:
        for row_offset, entry in enumerate(unassigned):
            row = ua_header_row + 1 + row_offset
            ws.cell(row=row, column=1, value=entry["indicator_id"])
            ws.cell(row=row, column=2, value=entry["indicator_name"]).alignment = Alignment(wrap_text=True)
            ws.cell(row=row, column=3, value=entry["working_group"])
            for i, abbrev in enumerate(CAMPUS_ORDER):
                cell = ws.cell(row=row, column=4 + i)
                cell.alignment = Alignment(horizontal="center")
                if abbrev in entry["missing"]:
                    cell.value = "Missing"
                    cell.fill = missing_fill
                    cell.font = missing_font
                else:
                    cell.value = "OK"
                    cell.fill = CAMPUS_BG_FILLS[abbrev]

        ua_end_row = ua_header_row + len(unassigned)
    else:
        ws.cell(row=ua_header_row + 1, column=1, value="All indicators have people assigned.")
        ws.cell(row=ua_header_row + 1, column=1).font = Font(italic=True, color="999999")
        ua_end_row = ua_header_row + 1

    # Excel Table for unassigned
    ua_table_ref = f"A{ua_header_row}:F{ua_end_row}"
    ua_table = Table(displayName="UnassignedYSE", ref=ua_table_ref)
    ua_table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2", showFirstColumn=False,
        showLastColumn=False, showRowStripes=True, showColumnStripes=False,
    )
    ws.add_table(ua_table)

    _auto_fit_columns(ws, max_width=50)
    ws.column_dimensions["A"].width = 8


def _build_comparison_sheet(wb: Workbook, wg_name: str, goals: list):
    """Build a working-group comparison sheet with one Excel Table per goal."""
    ws = wb.create_sheet(wg_name)

    # Title row
    ws.cell(row=1, column=1, value=f"{wg_name} — Campus Comparison")
    ws.cell(row=1, column=1).font = Font(bold=True, size=14)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=NUM_COLS)

    current_row = 3

    for goal in goals:
        current_row = _write_goal_table(ws, goal, current_row, wg_name)

    ws.freeze_panes = "A3"
    _auto_fit_columns(ws)
    ws.column_dimensions["A"].width = 6  # Indicator IDs are short (e.g. "12.34")


def _write_goal_table(ws, goal: dict, start_row: int, wg_name: str) -> int:
    """Write a single goal as an Excel Table with section headers above it."""
    goal_name = goal["goal_name"]
    goal_number = goal["goal_number"]
    indicators = goal["indicators"]
    current_row = start_row

    # --- Goal banner row (dark blue, merged) ---
    banner_text = f"Goal {goal_number}: {goal_name}"
    ws.cell(row=current_row, column=1, value=banner_text)
    for col_idx in range(1, NUM_COLS + 1):
        cell = ws.cell(row=current_row, column=col_idx)
        cell.font = GOAL_BANNER_FONT
        cell.fill = GOAL_BANNER_FILL
        cell.border = THIN_BORDER
    ws.merge_cells(
        start_row=current_row, start_column=1,
        end_row=current_row, end_column=NUM_COLS,
    )
    current_row += 1

    # --- Section header row (merged subsection labels) ---
    # Cols 1-2: blank but styled, col 2 gets black right border
    for c in range(1, 3):
        cell = ws.cell(row=current_row, column=c)
        cell.fill = SECTION_HEADER_FILL
        cell.border = THIN_BORDER if c == 1 else Border(
            left=_GREY_SIDE, right=_BLACK_SIDE, top=_GREY_SIDE, bottom=_GREY_SIDE)

    for col_start, col_end, label in SECTION_HEADERS:
        cell = ws.cell(row=current_row, column=col_start, value=label)
        cell.font = SECTION_HEADER_FONT
        cell.fill = SECTION_HEADER_FILL
        cell.alignment = HEADER_ALIGNMENT
        # Section header gets black border on right edge
        cell.border = Border(left=_BLACK_SIDE, right=_GREY_SIDE, top=_GREY_SIDE, bottom=_GREY_SIDE)
        if col_start != col_end:
            ws.merge_cells(
                start_row=current_row, start_column=col_start,
                end_row=current_row, end_column=col_end,
            )
        for c in range(col_start + 1, col_end + 1):
            cell2 = ws.cell(row=current_row, column=c)
            cell2.fill = SECTION_HEADER_FILL
            if c == col_end:
                cell2.border = Border(left=_GREY_SIDE, right=_BLACK_SIDE, top=_GREY_SIDE, bottom=_GREY_SIDE)
            else:
                cell2.border = THIN_BORDER
    current_row += 1

    # --- Column header row (this is the Excel Table header) ---
    # Campus columns get their campus color; cols 1-2 get the default blue
    table_header_row = current_row
    for col_idx, col_name in enumerate(COLUMNS, 1):
        cell = ws.cell(row=current_row, column=col_idx, value=col_name)
        cell.alignment = HEADER_ALIGNMENT
        if col_idx >= 3:
            cell.border = _campus_border(col_idx)
        elif col_idx == 2:
            cell.border = Border(left=_ROW_SIDE, right=_BLACK_SIDE, top=_ROW_SIDE, bottom=_ROW_SIDE)
        else:
            cell.border = Border(left=_ROW_SIDE, right=_ROW_SIDE, top=_ROW_SIDE, bottom=_ROW_SIDE)

        campus_i = COL_CAMPUS_INDEX.get(col_idx)
        if campus_i is not None:
            abbrev = CAMPUS_ORDER[campus_i]
            cell.fill = CAMPUS_HEADER_FILLS[abbrev]
            cell.font = CAMPUS_HEADER_FONTS[abbrev]
        else:
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
    current_row += 1

    # --- Data rows ---
    if not indicators:
        ws.cell(row=current_row, column=1, value="No data")
        ws.cell(row=current_row, column=1).font = Font(italic=True, color="999999")
        for col_idx in range(1, NUM_COLS + 1):
            cell = ws.cell(row=current_row, column=col_idx)
            if col_idx >= 3:
                cell.border = _campus_border(col_idx)
            elif col_idx == 2:
                cell.border = Border(left=_ROW_SIDE, right=_BLACK_SIDE, top=_ROW_SIDE, bottom=_ROW_SIDE)
            else:
                cell.border = Border(left=_ROW_SIDE, right=_ROW_SIDE, top=_ROW_SIDE, bottom=_ROW_SIDE)
            campus_i = COL_CAMPUS_INDEX.get(col_idx)
            if campus_i is not None:
                cell.fill = CAMPUS_BG_FILLS[CAMPUS_ORDER[campus_i]]
        current_row += 1
    else:
        for ind in indicators:
            current_row = _write_indicator_row(ws, ind, current_row)

    # --- Create Excel Table ---
    table_end_row = current_row - 1
    end_col_letter = get_column_letter(NUM_COLS)
    table_ref = f"A{table_header_row}:{end_col_letter}{table_end_row}"

    safe_wg = wg_name.replace(" ", "")[:8]
    safe_goal = str(goal_number)
    table_name = _next_table_name(f"{safe_wg}_G{safe_goal}_")

    table = Table(displayName=table_name, ref=table_ref)
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleLight9", showFirstColumn=False,
        showLastColumn=False, showRowStripes=True, showColumnStripes=False,
    )
    ws.add_table(table)

    # Bottom divider
    for col_idx in range(1, NUM_COLS + 1):
        cell = ws.cell(row=table_end_row, column=col_idx)
        cell.border = Border(
            left=cell.border.left, right=cell.border.right,
            top=cell.border.top,
            bottom=Side(style="medium", color="2F5496"),
        )

    # Spacer row
    for col_idx in range(1, NUM_COLS + 1):
        ws.cell(row=current_row, column=col_idx).fill = SPACER_FILL
    ws.row_dimensions[current_row].height = 8
    current_row += 2

    return current_row


def _write_indicator_row(ws, indicator: dict, row_num: int) -> int:
    """Write one indicator row across all 11 columns with campus colors and borders."""
    campuses = indicator["campuses"]

    # Col 1: indicator_id
    cell = ws.cell(row=row_num, column=1, value=indicator["indicator_id"])
    cell.border = Border(left=_ROW_SIDE, right=_ROW_SIDE, top=_ROW_SIDE, bottom=_ROW_SIDE)
    cell.alignment = Alignment(vertical="top")

    # Col 2: indicator_name — medium-weight right border to match section separators
    cell = ws.cell(row=row_num, column=2, value=indicator["indicator_name"])
    cell.border = Border(left=_ROW_SIDE, right=_BLACK_SIDE, top=_ROW_SIDE, bottom=_ROW_SIDE)
    cell.alignment = Alignment(vertical="top", wrap_text=True)

    # Cols 3-5: Status per campus
    for i, abbrev in enumerate(CAMPUS_ORDER):
        col = STATUS_COL_START + i
        status = campuses[abbrev]["status"]
        cell = ws.cell(row=row_num, column=col, value=status)
        cell.border = _campus_border(col)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        if status in STATUS_FILLS:
            cell.fill = STATUS_FILLS[status]
            cell.font = STATUS_FONTS[status]
        else:
            cell.fill = CAMPUS_BG_FILLS[abbrev]

    # Cols 6-8: Impl & Notes per campus
    for i, abbrev in enumerate(CAMPUS_ORDER):
        col = IMPL_COL_START + i
        cell = ws.cell(row=row_num, column=col, value=campuses[abbrev]["impl_and_notes"])
        cell.border = _campus_border(col)
        cell.alignment = Alignment(vertical="top", wrap_text=True)
        cell.fill = CAMPUS_BG_FILLS[abbrev]

    # Cols 9-11: People & Orgs per campus
    for i, abbrev in enumerate(CAMPUS_ORDER):
        col = PEOPLE_COL_START + i
        cell = ws.cell(row=row_num, column=col, value=campuses[abbrev]["people_and_orgs"])
        cell.border = _campus_border(col)
        cell.alignment = Alignment(vertical="top", wrap_text=True)
        cell.fill = CAMPUS_BG_FILLS[abbrev]

    return row_num + 1


def _auto_fit_columns(ws, min_width=10, max_width=50):
    """Set column widths based on maximum content length."""
    for col_cells in ws.columns:
        max_length = 0
        col_letter = get_column_letter(col_cells[0].column)
        for cell in col_cells:
            if cell.value is not None:
                try:
                    cell_len = len(str(cell.value))
                    if cell_len > max_length:
                        max_length = cell_len
                except Exception:
                    pass
        adjusted = max(min(max_length + 2, max_width), min_width)
        ws.column_dimensions[col_letter].width = adjusted
